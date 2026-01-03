from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
import tempfile
import os
from PySide6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
def export_to_excel_from_tableview(QTableView,path=""):
    model = QTableView.model()
    if not model:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)

    # Ghi tiêu đề
    for col in range(model.columnCount()):
        header = model.headerData(col, Qt.Horizontal)
        cell = ws.cell(row=1, column=col + 1, value=header)
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.font = bold_font

    # Ghi dữ liệu
    for row in range(model.rowCount()):
        for col in range(model.columnCount()):
            index = model.index(row, col)
            data = index.data()
            cell = ws.cell(row=row + 2, column=col + 1, value=data)
            cell.alignment = center_alignment  # 👈 Căn giữa ô
            cell.number_format = '@'  # 👈 Định dạng text để tránh lỗi số bị đổi

    # 3. Auto width cho từng cột
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value = str(cell.value)
                max_length = max(max_length, len(value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    if not path:
        # Tạo file tạm và mở
        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp_file.name)
        os.startfile(tmp_file.name)  # 👈 Mở Excel ngay sau khi tạo
        return True
    else:
        # Lưu file
        try:
            wb.save(path)
            #os.startfile(path)  # 👈 Mở file lên luôn nếu bạn muốn
            return True
        except Exception as e:
            return False
    return False
